from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, func, or_
from typing import Optional
from datetime import date
from math import ceil
import io

from app.database import get_db
from app.models import Proceso, Entidad, CatEstado, CatTipoProceso, CatDistrito, CatProvincia, CatRegion
from app.schemas import ProcesosPaginados, ProcesoListItem, ProcesoDetalle, Paginacion

router = APIRouter(prefix="/procesos", tags=["procesos"])


def _build_query(db, entidad, estado, tipo, nivel, departamento, provincia,
                 monto_min, monto_max, fecha_desde, fecha_hasta, texto, ubigeo):
    q = (
        db.query(
            Proceso.id,
            Proceso.codigo_seace,
            Entidad.nombre.label("entidad"),
            Proceso.objeto_contratacion,
            CatTipoProceso.nombre.label("tipo_proceso"),
            CatEstado.nombre.label("estado"),
            Proceso.valor_referencial,
            Proceso.moneda,
            Proceso.fecha_convocatoria,
            Proceso.nivel_gobierno,
            Proceso.entidad_departamento.label("departamento"),
            Proceso.entidad_provincia.label("provincia"),
            Proceso.url_seace,
        )
        .join(Entidad, Proceso.entidad_id == Entidad.id)
        .outerjoin(CatTipoProceso, Proceso.tipo_proceso_id == CatTipoProceso.id)
        .outerjoin(CatEstado, Proceso.estado_id == CatEstado.id)
    )
    if entidad:
        q = q.filter(Entidad.nombre_norm.ilike(f"%{entidad}%"))
    if estado:
        q = q.filter(CatEstado.codigo == estado)
    if tipo:
        q = q.filter(CatTipoProceso.codigo == tipo)
    if nivel:
        q = q.filter(Proceso.nivel_gobierno == nivel)
    if departamento:
        q = q.filter(Proceso.entidad_departamento.ilike(f"%{departamento}%"))
    if provincia:
        q = q.filter(Proceso.entidad_provincia.ilike(f"%{provincia}%"))
    if monto_min is not None:
        q = q.filter(Proceso.valor_referencial >= monto_min)
    if monto_max is not None:
        q = q.filter(Proceso.valor_referencial <= monto_max)
    if fecha_desde:
        q = q.filter(Proceso.fecha_convocatoria >= fecha_desde)
    if fecha_hasta:
        q = q.filter(Proceso.fecha_convocatoria <= fecha_hasta)
    if ubigeo:
        q = q.filter(Proceso.ubigeo.startswith(ubigeo))
    if texto:
        q = q.filter(Proceso.objeto_contratacion.ilike(f"%{texto}%"))
    return q


@router.get("", response_model=ProcesosPaginados)
def listar_procesos(
    entidad: Optional[str] = None,
    estado: Optional[str] = None,
    tipo: Optional[str] = None,
    nivel: Optional[str] = None,
    departamento: Optional[str] = None,
    provincia: Optional[str] = None,
    monto_min: Optional[float] = None,
    monto_max: Optional[float] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    texto: Optional[str] = None,
    ubigeo: Optional[str] = None,
    pagina: int = Query(default=1, ge=1),
    por_pagina: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    q = _build_query(db, entidad, estado, tipo, nivel, departamento, provincia,
                     monto_min, monto_max, fecha_desde, fecha_hasta, texto, ubigeo)
    total = q.count()
    rows = q.order_by(Proceso.fecha_convocatoria.desc()).offset((pagina - 1) * por_pagina).limit(por_pagina).all()
    items = [ProcesoListItem(**r._asdict()) for r in rows]
    return ProcesosPaginados(
        data=items,
        paginacion=Paginacion(total=total, pagina=pagina, por_pagina=por_pagina, paginas=ceil(total / por_pagina) if total else 0),
    )


@router.get("/exportar")
def exportar_procesos(
    entidad: Optional[str] = None,
    estado: Optional[str] = None,
    tipo: Optional[str] = None,
    nivel: Optional[str] = None,
    departamento: Optional[str] = None,
    provincia: Optional[str] = None,
    monto_min: Optional[float] = None,
    monto_max: Optional[float] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    texto: Optional[str] = None,
    ubigeo: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Exporta los procesos filtrados a un archivo Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = _build_query(db, entidad, estado, tipo, nivel, departamento, provincia,
                     monto_min, monto_max, fecha_desde, fecha_hasta, texto, ubigeo)
    rows = q.order_by(Proceso.fecha_convocatoria.desc()).limit(10_000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procesos SEACE"

    # ── Estilos ──────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1D4ED8")   # azul oscuro
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align   = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Encabezados ──────────────────────────────────────────────────────────
    COLS = [
        ("Código SEACE",    18),
        ("Entidad",         40),
        ("Objeto / Servicio / Obra", 55),
        ("Tipo de Proceso", 22),
        ("Estado",          18),
        ("Nivel de Gobierno", 22),
        ("Departamento",    18),
        ("Provincia",       18),
        ("Monto (S/)",      15),
        ("Moneda",          10),
        ("Fecha Convocatoria", 18),
        ("URL SEACE",       40),
    ]

    for col_idx, (titulo, ancho) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Datos ────────────────────────────────────────────────────────────────
    fill_par  = PatternFill("solid", fgColor="EFF6FF")   # azul muy claro filas pares
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, r in enumerate(rows, start=2):
        fila = [
            r.codigo_seace,
            r.entidad,
            r.objeto_contratacion,
            r.tipo_proceso,
            r.estado,
            r.nivel_gobierno,
            r.departamento,
            r.provincia,
            r.valor_referencial,
            r.moneda,
            str(r.fecha_convocatoria) if r.fecha_convocatoria else None,
            r.url_seace,
        ]
        fill = fill_par if row_idx % 2 == 0 else fill_impar
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = cell_align
            cell.border = border
            cell.fill = fill
            # Monto: formato numérico
            if col_idx == 9 and valor is not None:
                cell.number_format = '#,##0.00'

    # ── Autofiltro ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # ── Stream ───────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "procesos_seace.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Access-Control-Expose-Headers": "Content-Disposition",
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/{codigo_seace}", response_model=ProcesoDetalle)
def detalle_proceso(codigo_seace: str, db: Session = Depends(get_db)):
    row = (
        db.query(
            Proceso.id, Proceso.codigo_seace, Proceso.nro_expediente,
            Entidad.nombre.label("entidad"),
            Proceso.objeto_contratacion, Proceso.descripcion,
            CatTipoProceso.nombre.label("tipo_proceso"),
            CatEstado.nombre.label("estado"),
            Proceso.valor_referencial, Proceso.moneda,
            Proceso.fecha_convocatoria, Proceso.fecha_buena_pro, Proceso.fecha_suscripcion,
            Proceso.nivel_gobierno,
            Proceso.entidad_departamento.label("departamento"),
            Proceso.entidad_provincia.label("provincia"),
            Proceso.entidad_distrito.label("distrito"),
            Proceso.ubigeo, Proceso.url_seace, Proceso.ingested_at,
        )
        .join(Entidad, Proceso.entidad_id == Entidad.id)
        .outerjoin(CatTipoProceso, Proceso.tipo_proceso_id == CatTipoProceso.id)
        .outerjoin(CatEstado, Proceso.estado_id == CatEstado.id)
        .filter(Proceso.codigo_seace == codigo_seace)
        .first()
    )
    if not row:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Proceso no encontrado")
    return ProcesoDetalle(**row._asdict())
